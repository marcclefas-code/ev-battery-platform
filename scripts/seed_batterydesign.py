import asyncio
import uuid
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
from sqlalchemy import select, text
from app.services.database import get_db_session
from app.models.entity import BatteryEntity
from app.models.part_number import BatteryPartNumber
from app.models.property_definition import PropertyDefinition
from app.models.property_statement import PropertyStatement
from app.models.vehicle_variant_code import VehicleVariantCode
import structlog

logger = structlog.get_logger()

BDN_SOURCE = "batterydesign.net"
BDN_SOURCE_URL = "https://batterydesign.net"

DATA_DIR = os.environ.get("BDN_DATA_DIR", r"C:\EVscraper\batterydesigndotnet")

CELL_FILE = os.path.join(DATA_DIR, "cell-database-BatteryDesignNET-v1.4.01-download.xlsx")
MODULE_FILE = os.path.join(DATA_DIR, "module-benchmark-specification-BatteryDesignNET-v1.2.08-download.xlsx")
PACK_FILE = os.path.join(DATA_DIR, "pack-benchmark-specification-BatteryDesignNET-v1.5-download.xlsx")
VEHICLE_FILE = os.path.join(DATA_DIR, "vehicle-model-0.25.xlsx")
MANUFACTURER_FILE = os.path.join(DATA_DIR, "Cell-Manufacturers-0.107.xlsx")

PROPERTY_DEFINITIONS = [
    ("bdn_ref", "BatteryDesign.NET Reference", None, "string", ["hv_cell", "hv_module", "pack"], "identifier"),
    ("cell_name", "Cell Name", None, "string", ["hv_cell"], "identification"),
    ("manufacturer", "Manufacturer", None, "string", ["hv_cell", "hv_module", "pack"], "identification"),
    ("form_factor", "Form Factor", None, "string", ["hv_cell", "hv_module", "pack"], "physical"),
    ("capacity_ah_nom", "Nominal Capacity", "Ah", "float", ["hv_cell", "hv_module", "pack"], "electrical"),
    ("capacity_ah_max", "Maximum Capacity", "Ah", "float", ["hv_cell"], "electrical"),
    ("capacity_ah_min", "Minimum Capacity", "Ah", "float", ["hv_cell"], "electrical"),
    ("voltage_nom", "Nominal Voltage", "V", "float", ["hv_cell", "hv_module", "pack"], "electrical"),
    ("voltage_max", "Maximum Voltage", "V", "float", ["hv_cell", "hv_module", "pack"], "electrical"),
    ("voltage_min", "Minimum Voltage", "V", "float", ["hv_cell", "hv_module", "pack"], "electrical"),
    ("energy_wh", "Energy", "Wh", "float", ["hv_cell", "hv_module", "pack"], "electrical"),
    ("weight_kg", "Weight", "kg", "float", ["hv_cell", "hv_module", "pack"], "physical"),
    ("cathode_chemistry", "Cathode Chemistry", None, "string", ["hv_cell"], "chemistry"),
    ("cathode_type", "Cathode Type", None, "string", ["hv_cell"], "chemistry"),
    ("anode_chemistry", "Anode Chemistry", None, "string", ["hv_cell"], "chemistry"),
    ("electrode_layout", "Electrode Layout", None, "string", ["hv_cell"], "physical"),
    ("tab_type", "Tab Type", None, "string", ["hv_cell"], "physical"),
    ("internal_resistance", "Internal Resistance", "mΩ", "float", ["hv_cell"], "electrical"),
    ("primary_application", "Primary Application", None, "string", ["hv_module", "pack"], "identification"),
    ("soc_range_recommended", "Recommended SoC Range", "%", "string", ["hv_module", "pack"], "electrical"),
    ("module_dcir", "Module DCIR", "mΩ", "float", ["hv_module"], "electrical"),
    ("terminals", "Terminals", None, "string", ["hv_module", "pack"], "physical"),
    ("fuse", "Fuse", None, "string", ["hv_module"], "protection"),
    ("power", "Power", "kW", "float", ["hv_module", "pack"], "electrical"),
    ("manufacturers_code", "Manufacturers Code Number", None, "string", ["pack"], "identification"),
    ("application", "Application", None, "string", ["pack"], "identification"),
    ("first_produced", "First Produced", None, "string", ["pack"], "identification"),
    ("soc_window", "SoC Window", "%", "string", ["pack"], "electrical"),
    ("voltage_range", "Voltage Range", "V", "string", ["pack"], "electrical"),
    ("num_modules", "Number of Modules", None, "integer", ["pack"], "configuration"),
    ("num_cells", "Number of Cells", None, "integer", ["hv_module", "pack"], "configuration"),
    ("battery_type", "Battery Type", None, "string", ["hv_cell", "hv_module", "pack"], "chemistry"),
    ("tesla_model", "Tesla Model", None, "string", ["pack"], "identification"),
    ("steady_state_consumption", "Steady State Consumption", "kWh/100km", "float", ["pack"], "performance"),
    ("ev_make", "EV Make", None, "string", ["pack"], "vehicle"),
    ("ev_model", "EV Model", None, "string", ["pack"], "vehicle"),
]


def normalize_pn(raw: str) -> str:
    return raw.strip().upper().replace(" ", "").replace("-", "") if raw else ""


async def seed_property_definitions():
    logger.info("Seeding property definitions...")
    async with get_db_session() as session:
        for code, label, unit, data_type, applies_to, category in PROPERTY_DEFINITIONS:
            result = await session.execute(
                select(PropertyDefinition).where(PropertyDefinition.code == code)
            )
            existing = result.scalar_one_or_none()
            if not existing:
                pd = PropertyDefinition(
                    code=code,
                    label=label,
                    unit=unit,
                    data_type=data_type,
                    applies_to=applies_to,
                    category=category,
                )
                session.add(pd)
        await session.commit()
    logger.info("Property definitions seeded")


async def seed_cells(file_path: str) -> int:
    logger.info("Seeding cells from %s", file_path)
    if not os.path.exists(file_path):
        logger.warning("File not found: %s", file_path)
        return 0

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb["Cells"]
    rows = list(ws.iter_rows(min_row=3, values_only=True))
    wb.close()

    count = 0
    async with get_db_session() as session:
        for row in rows:
            if not row or not row[3]:
                continue

            try:
                bd_ref = str(row[1]).strip() if row[1] else None
                manufacturer = str(row[2]).strip() if row[2] else None
                cell_name = str(row[3]).strip() if row[3] else None
                if not bd_ref or not cell_name:
                    continue

                entity_id = uuid.uuid4()
                entity = BatteryEntity(
                    id=entity_id,
                    entity_type="hv_cell",
                    canonical_name=f"{manufacturer} {cell_name}" if manufacturer else cell_name,
                    normalized_primary_part_number=f"BDN:{bd_ref}",
                    occurrence_count=1,
                    status="confirmed",
                )
                session.add(entity)

                pn = BatteryPartNumber(
                    id=uuid.uuid4(),
                    entity_id=entity_id,
                    raw=bd_ref,
                    normalized=normalize_pn(bd_ref),
                    brand="batterydesign.net",
                    pn_type="service",
                    source_url=BDN_SOURCE_URL,
                    evidence_quote=f"BatteryDesign.NET cell database ref {bd_ref}",
                    trust_weight=0.85,
                )
                session.add(pn)

                props = [
                    ("bdn_ref", bd_ref, None),
                    ("cell_name", cell_name, None),
                    ("manufacturer", manufacturer, None),
                    ("capacity_ah_nom", row[4], "Ah"),
                    ("capacity_ah_max", row[5], "Ah"),
                    ("capacity_ah_min", row[6], "Ah"),
                    ("voltage_max", row[7], "V"),
                    ("voltage_nom", row[8], "V"),
                    ("voltage_min", row[9], "V"),
                    ("energy_wh", row[10], "Wh"),
                    ("weight_kg", row[13], "kg"),
                    ("form_factor", row[14], None),
                    ("cathode_chemistry", row[15], None),
                    ("cathode_type", row[16], None),
                    ("electrode_layout", row[24], None),
                    ("tab_type", row[25], None),
                    ("internal_resistance", row[26], "mΩ"),
                ]

                for code, val, unit in props:
                    if val is not None and str(val).strip() not in ("", "-", "None"):
                        try:
                            stmt = PropertyStatement(
                                id=uuid.uuid4(),
                                entity_id=entity_id,
                                code=code,
                                value={"value": float(val)} if unit or code != "bdn_ref" else {"value": str(val)},
                                unit=unit,
                                source_url=BDN_SOURCE_URL,
                                trust_weight=0.85,
                                confidence=0.90,
                                status="confirmed",
                            )
                            session.add(stmt)
                        except (ValueError, TypeError):
                            pass

                count += 1
                if count % 100 == 0:
                    await session.commit()
                    logger.info("  Committed %d cells...", count)

            except Exception as e:
                logger.warning("Error processing cell row: %s", e)
                continue

        await session.commit()
    logger.info("Seeded %d cells", count)
    return count


async def seed_modules(file_path: str) -> int:
    logger.info("Seeding modules from %s", file_path)
    if not os.path.exists(file_path):
        logger.warning("File not found: %s", file_path)
        return 0

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb["Modules"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    count = 0
    async with get_db_session() as session:
        for row in rows:
            if not row or not row[1] or str(row[1]).strip() in ("", "database ref", "-"):
                continue
            try:
                bd_ref = str(row[1]).strip() if row[1] else None
                manufacturer = str(row[2]).strip() if row[2] else None
                module_name = str(row[3]).strip() if row[3] else None
                if not bd_ref:
                    continue

                entity_id = uuid.uuid4()
                entity = BatteryEntity(
                    id=entity_id,
                    entity_type="hv_module",
                    canonical_name=f"{manufacturer} {module_name}" if manufacturer else module_name,
                    normalized_primary_part_number=f"BDN:MOD:{bd_ref}",
                    occurrence_count=1,
                    status="confirmed",
                )
                session.add(entity)

                pn = BatteryPartNumber(
                    id=uuid.uuid4(),
                    entity_id=entity_id,
                    raw=bd_ref,
                    normalized=normalize_pn(bd_ref),
                    brand="batterydesign.net",
                    pn_type="service",
                    source_url=BDN_SOURCE_URL,
                    evidence_quote=f"BatteryDesign.NET module database ref {bd_ref}",
                    trust_weight=0.85,
                )
                session.add(pn)

                props = [
                    ("bdn_ref", bd_ref, None),
                    ("manufacturer", manufacturer, None),
                    ("module_name", module_name, None),
                    ("primary_application", row[4], None),
                    ("capacity_ah_nom", row[5], "Ah"),
                    ("voltage_nom", row[6], "V"),
                    ("voltage_max", row[7], "V"),
                    ("voltage_min", row[8], "V"),
                    ("energy_wh", row[9], "Wh"),
                    ("soc_range_recommended", row[11], "%"),
                    ("module_dcir", row[13], "mΩ"),
                    ("terminals", row[15], None),
                    ("fuse", row[17], None),
                    ("power", row[19], "kW"),
                ]

                for code, val, unit in props:
                    if val is not None and str(val).strip() not in ("", "-", "None"):
                        try:
                            stmt = PropertyStatement(
                                id=uuid.uuid4(),
                                entity_id=entity_id,
                                code=code,
                                value={"value": float(val)} if unit else {"value": str(val)},
                                unit=unit,
                                source_url=BDN_SOURCE_URL,
                                trust_weight=0.85,
                                confidence=0.90,
                                status="confirmed",
                            )
                            session.add(stmt)
                        except (ValueError, TypeError):
                            pass

                count += 1
                if count % 50 == 0:
                    await session.commit()
                    logger.info("  Committed %d modules...", count)

            except Exception as e:
                logger.warning("Error processing module row: %s", e)
                continue

        await session.commit()
    logger.info("Seeded %d modules", count)
    return count


async def seed_packs(file_path: str) -> int:
    logger.info("Seeding packs from %s", file_path)
    if not os.path.exists(file_path):
        logger.warning("File not found: %s", file_path)
        return 0

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb["Packs"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    count = 0
    async with get_db_session() as session:
        for row in rows:
            if not row or not row[2] or str(row[2]).strip() in ("", "database ref", "-"):
                continue
            try:
                bd_ref = str(row[2]).strip() if row[2] else None
                manufacturer = str(row[3]).strip() if row[3] else None
                model_pack = str(row[4]).strip() if row[4] else None
                manufacturers_code = str(row[5]).strip() if row[5] else None
                application = str(row[6]).strip() if row[6] else None
                if not bd_ref:
                    continue

                entity_id = uuid.uuid4()
                entity = BatteryEntity(
                    id=entity_id,
                    entity_type="pack",
                    canonical_name=f"{manufacturer} {model_pack}" if manufacturer else model_pack,
                    normalized_primary_part_number=f"BDN:PACK:{bd_ref}",
                    occurrence_count=1,
                    status="confirmed",
                )
                session.add(entity)

                pn = BatteryPartNumber(
                    id=uuid.uuid4(),
                    entity_id=entity_id,
                    raw=bd_ref,
                    normalized=normalize_pn(bd_ref),
                    brand="batterydesign.net",
                    pn_type="service",
                    source_url=BDN_SOURCE_URL,
                    evidence_quote=f"BatteryDesign.NET pack database ref {bd_ref}",
                    trust_weight=0.85,
                )
                session.add(pn)

                props = [
                    ("bdn_ref", bd_ref, None),
                    ("manufacturer", manufacturer, None),
                    ("tesla_model", model_pack, None),
                    ("manufacturers_code", manufacturers_code, None),
                    ("application", application, None),
                    ("first_produced", row[7], None),
                    ("energy_wh", row[8], "Wh"),
                    ("soc_window", row[10], "%"),
                    ("voltage_nom", row[11], "V"),
                    ("voltage_range", row[12], "V"),
                    ("voltage_max", row[13], "V"),
                    ("voltage_min", row[14], "V"),
                    ("capacity_ah_nom", row[15], "Ah"),
                    ("weight_kg", row[16], "kg"),
                    ("power", row[17], "kW"),
                    ("num_modules", row[18], None),
                    ("num_cells", row[19], None),
                    ("soc_range_recommended", row[20], "%"),
                    ("form_factor", row[21], None),
                    ("battery_type", row[22], None),
                ]

                for code, val, unit in props:
                    if val is not None and str(val).strip() not in ("", "-", "None"):
                        try:
                            stmt = PropertyStatement(
                                id=uuid.uuid4(),
                                entity_id=entity_id,
                                code=code,
                                value={"value": float(val)} if unit else {"value": str(val)},
                                unit=unit,
                                source_url=BDN_SOURCE_URL,
                                trust_weight=0.85,
                                confidence=0.90,
                                status="confirmed",
                            )
                            session.add(stmt)
                        except (ValueError, TypeError):
                            pass

                count += 1
                if count % 200 == 0:
                    await session.commit()
                    logger.info("  Committed %d packs...", count)

            except Exception as e:
                logger.warning("Error processing pack row: %s", e)
                continue

        await session.commit()
    logger.info("Seeded %d packs", count)
    return count


async def seed_vehicles(file_path: str) -> int:
    logger.info("Seeding vehicle fitment from %s", file_path)
    if not os.path.exists(file_path):
        logger.warning("File not found: %s", file_path)
        return 0

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb["VehData"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    count = 0
    async with get_db_session() as session:
        for row in rows:
            if not row or not row[2]:
                continue
            try:
                make = str(row[2]).strip() if row[2] else None
                model = str(row[3]).strip() if row[3] else None
                if not make or not model or make == "-" or model == "-":
                    continue

                result = await session.execute(
                    select(BatteryEntity).where(
                        BatteryEntity.canonical_name.ilike(f"%{model}%")
                    ).limit(1)
                )
                entity = result.scalar_one_or_none()
                if entity:
                    vvc = VehicleVariantCode(
                        id=uuid.uuid4(),
                        entity_id=entity.id,
                        make=make,
                        model=model,
                    )
                    session.add(vvc)
                    count += 1

            except Exception as e:
                logger.warning("Error processing vehicle row: %s", e)
                continue

        await session.commit()
    logger.info("Seeded %d vehicle fitment records", count)
    return count


async def seed_manufacturers(file_path: str) -> int:
    logger.info("Seeding manufacturers from %s", file_path)
    if not os.path.exists(file_path):
        logger.warning("File not found: %s", file_path)
        return 0

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb["Sheet1"]
    rows = list(ws.iter_rows(min_row=3, values_only=True))
    wb.close()

    count = 0
    async with get_db_session() as session:
        for row in rows:
            if not row or not row[0] or str(row[0]).strip() in ("", "Manufacturer"):
                continue
            try:
                mfr_name = str(row[0]).strip() if row[0] else None
                if not mfr_name:
                    continue

                entity_id = uuid.uuid4()
                entity = BatteryEntity(
                    id=entity_id,
                    entity_type="hv_cell",
                    canonical_name=mfr_name,
                    normalized_primary_part_number=f"BDN:MFR:{normalize_pn(mfr_name)}",
                    occurrence_count=1,
                    status="confirmed",
                )
                session.add(entity)

                pn = BatteryPartNumber(
                    id=uuid.uuid4(),
                    entity_id=entity_id,
                    raw=mfr_name,
                    normalized=normalize_pn(mfr_name),
                    brand="batterydesign.net",
                    pn_type="service",
                    source_url=BDN_SOURCE_URL,
                    evidence_quote=f"BatteryDesign.NET manufacturer: {mfr_name}",
                    trust_weight=0.85,
                )
                session.add(pn)

                props = [
                    ("manufacturer", mfr_name, None),
                    ("bdn_ref", row[0], None),
                    ("description", row[1], None),
                    ("cell_formats", row[4], None),
                    ("cathode_chemistry", row[5], None),
                    ("products", row[6], None),
                    ("hq", row[7], None),
                    ("production_2020", row[8], "GWh"),
                    ("production_2021", row[9], "GWh"),
                    ("production_2022", row[10], "GWh"),
                    ("production_2023", row[11], "GWh"),
                    ("production_2024", row[12], "GWh"),
                    ("production_2025", row[13], "GWh"),
                    ("production_2026", row[14], "GWh"),
                ]

                for code, val, unit in props:
                    if val is not None and str(val).strip() not in ("", "-", "None"):
                        try:
                            stmt = PropertyStatement(
                                id=uuid.uuid4(),
                                entity_id=entity_id,
                                code=code,
                                value={"value": float(val)} if unit else {"value": str(val)},
                                unit=unit,
                                source_url=BDN_SOURCE_URL,
                                trust_weight=0.80,
                                confidence=0.80,
                                status="confirmed",
                            )
                            session.add(stmt)
                        except (ValueError, TypeError):
                            pass

                count += 1

            except Exception as e:
                logger.warning("Error processing manufacturer row: %s", e)
                continue

        await session.commit()
    logger.info("Seeded %d manufacturers", count)
    return count


async def main():
    logger.info("=== BatteryDesign.NET Seed Script ===")
    logger.info("Data source: %s", os.path.dirname(CELL_FILE))

    await seed_property_definitions()

    cell_count = await seed_cells(CELL_FILE)
    module_count = await seed_modules(MODULE_FILE)
    pack_count = await seed_packs(PACK_FILE)
    mfr_count = await seed_manufacturers(MANUFACTURER_FILE)
    vehicle_count = await seed_vehicles(VEHICLE_FILE)

    logger.info("=== Seed Complete ===")
    logger.info("Total: %d cells, %d modules, %d packs, %d manufacturers, %d vehicle fitment",
                cell_count, module_count, pack_count, mfr_count, vehicle_count)


if __name__ == "__main__":
    asyncio.run(main())
